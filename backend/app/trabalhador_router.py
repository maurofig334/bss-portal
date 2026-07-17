"""
Endpoints de trabalhador.

GET /trabalhadores              → lista com paginação + filtros
GET /trabalhadores/exportar     → .xlsx com o resultado do filtro (sem paginar)
GET /trabalhadores/{id}         → detalhe
"""

from datetime import date, datetime
from io import BytesIO
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse

from .auth import UsuarioInfo, usuario_logado
from . import trabalhador_repo


router = APIRouter(prefix="/trabalhadores", tags=["trabalhadores"])


def _escopo_por_perfil(
    usuario: UsuarioInfo,
    id_empresa: int | None,
    id_sindicato: int | None,
) -> tuple[int | None, int | None] | None:
    """
    Aplica o escopo do perfil. Devolve (id_empresa, id_sindicato) ajustados,
    ou None quando o usuário não tem vínculo nenhum (chamador deve retornar
    lista vazia).

    Compartilhado entre listar() e exportar() — se a exportação tivesse a
    própria cópia dessa regra, uma hora as duas divergiriam e o .xlsx viraria
    o furo de RLS mais silencioso possível: o cliente baixa, abre no Excel, e
    ninguém nunca vê a tela que mostrou demais.
    """
    if usuario.perfil == "empresa":
        if not usuario.empresas:
            return None
        if id_empresa is None:
            id_empresa = usuario.empresas[0]
        elif id_empresa not in usuario.empresas:
            raise HTTPException(403, "Empresa fora do escopo do usuário")
    elif usuario.perfil == "sindicato":
        if not usuario.sindicatos:
            return None
        if id_sindicato is None:
            id_sindicato = usuario.sindicatos[0]
        elif id_sindicato not in usuario.sindicatos:
            raise HTTPException(403, "Sindicato fora do escopo do usuário")
    return id_empresa, id_sindicato


@router.get("")
def listar(
    usuario: Annotated[UsuarioInfo, Depends(usuario_logado)],
    busca: str | None = Query(None, description="Busca por nome ou CPF"),
    situacao: str | None = Query(None, description="ativo, inativo, carencia"),
    id_empresa: int | None = None,
    id_sindicato: int | None = None,
    uf: str | None = Query(None, max_length=2),
    pagina: int = 1,
    por_pagina: int = 50,
    ordem: str = "nome_completo",
    desc: bool = False,
):
    """
    Lista trabalhadores com filtros.
    Permissões: empresa só vê suas empresas; sindicato só vê seu sindicato.
    Admin/interno/analista veem tudo.
    """
    # RLS por perfil
    escopo = _escopo_por_perfil(usuario, id_empresa, id_sindicato)
    if escopo is None:
        return {"linhas": [], "total": 0, "pagina": 1,
                "por_pagina": por_pagina, "paginas": 0}
    id_empresa, id_sindicato = escopo

    return trabalhador_repo.listar(
        busca=busca,
        situacao=situacao,
        id_empresa=id_empresa,
        id_sindicato=id_sindicato,
        uf=uf,
        pagina=pagina,
        por_pagina=por_pagina,
        ordem=ordem,
        desc=desc,
    )


@router.get("/exportar")
def exportar(
    usuario: Annotated[UsuarioInfo, Depends(usuario_logado)],
    busca: str | None = None,
    situacao: str | None = Query(None, description="ativo, inativo, carencia"),
    id_empresa: int | None = None,
    id_sindicato: int | None = None,
    uf: str | None = Query(None, max_length=2),
    ordem: str = "nome_completo",
    desc: bool = False,
):
    """
    Exporta o resultado do filtro em .xlsx — TODAS as linhas, sem paginação.

    Recebe os mesmos parâmetros de GET /trabalhadores de propósito: a tela
    manda o filtro que está aplicado e o arquivo sai igual ao que está na tela.

    ATENÇÃO À ORDEM DAS ROTAS: esta precisa vir ANTES de /{id_trabalhador},
    senão o FastAPI casa "exportar" com aquela rota e tenta convertê-la em int
    (erro 422). Rota literal antes de rota com parâmetro.
    """
    escopo = _escopo_por_perfil(usuario, id_empresa, id_sindicato)
    if escopo is None:
        raise HTTPException(403, "Usuário sem vínculo — nada a exportar")
    id_empresa, id_sindicato = escopo

    linhas = trabalhador_repo.listar_tudo(
        busca=busca, situacao=situacao, id_empresa=id_empresa,
        id_sindicato=id_sindicato, uf=uf, ordem=ordem, desc=desc,
    )

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Trabalhadores"

    cabecalho = [
        "Tipo", "Nome completo", "CPF", "Situação",
        "CPF do titular", "Dependentes ativos",
        "Empresa", "CNPJ", "Sindicato",
        "Cidade", "UF", "Último vínculo", "Último pagamento",
    ]
    ws.append(cabecalho)

    bold = Font(bold=True)
    fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    for cell in ws[1]:
        cell.font = bold
        cell.fill = fill

    def _dt(v):
        # openpyxl escreve date/datetime nativos; o resto vai como texto.
        if isinstance(v, (date, datetime)):
            return v
        return v or ""

    for t in linhas:
        ws.append([
            "Dependente" if t["titularidade"] == "dependente" else "Trabalhador",
            t["nome_completo"] or "",
            t["cpf"] or "",
            t["situacao"] or "",
            t["cpf_titular"] or "",
            t["qtd_dependentes_ativos"] or 0,
            t["empresa"] or "",
            t["empresa_cnpj"] or "",
            t["sindicato"] or "",
            t["trab_cidade"] or "",
            t["trab_uf"] or "",
            _dt(t["mes_ultimo_vinculo"]),
            _dt(t["ultimo_pagamento_em"]),
        ])

    larguras = {"A": 12, "B": 38, "C": 14, "D": 11, "E": 14, "F": 8,
                "G": 38, "H": 20, "I": 38, "J": 20, "K": 5, "L": 14, "M": 16}
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"      # cabeçalho fixo ao rolar
    ws.auto_filter.ref = ws.dimensions

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    hoje = date.today().isoformat()
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="trabalhadores_{hoje}.xlsx"'},
    )


@router.get("/dependentes/{cpf_titular}")
def listar_dependentes(
    cpf_titular: str,
    usuario: Annotated[UsuarioInfo, Depends(usuario_logado)],
):
    """
    Retorna dependentes vinculados a um titular pelo CPF.

    RLS pelo TITULAR: os dependentes herdam a empresa dele, então é o titular
    que responde "esse CPF está no seu escopo?".

    Este endpoint recebia o `usuario` e nunca o usava: com um CPF qualquer,
    devolvia nome, CPF e data de nascimento dos dependentes de trabalhador de
    qualquer empresa. Era o único endpoint de trabalhador sem RLS — e o dado
    exposto é o mais sensível que temos (menores de idade, LGPD).
    """
    cpf = "".join(c for c in cpf_titular if c.isdigit())
    if len(cpf) != 11:
        raise HTTPException(400, "CPF inválido")

    if usuario.perfil in ("empresa", "sindicato"):
        titular = trabalhador_repo.buscar_titular(cpf)
        # 404 e não 403: quem não pode ver a empresa também não precisa
        # descobrir que o CPF existe na base.
        if not titular:
            raise HTTPException(404, "Titular não encontrado")
        if usuario.perfil == "empresa" and \
                titular.get("id_empresa_atual") not in usuario.empresas:
            raise HTTPException(404, "Titular não encontrado")
        if usuario.perfil == "sindicato" and \
                titular.get("id_sindicato_atual") not in usuario.sindicatos:
            raise HTTPException(404, "Titular não encontrado")

    return trabalhador_repo.buscar_dependentes(cpf)


@router.get("/{id_trabalhador}/detalhe")
def detalhe_completo(
    id_trabalhador: int,
    usuario: Annotated[UsuarioInfo, Depends(usuario_logado)],
):
    """
    Detalhe completo do trabalhador (#37): dados + endereço + dependentes
    relacionados (se titular) ou titular vinculado (se dependente).
    """
    row = trabalhador_repo.buscar_detalhe(id_trabalhador)
    if not row:
        raise HTTPException(404, "Trabalhador não encontrado")

    # RLS: empresa só vê trabalhador das suas empresas; sindicato só do seu sind
    if usuario.perfil == "empresa" and row.get("id_empresa_atual") not in usuario.empresas:
        raise HTTPException(403, "Trabalhador fora do escopo do usuário")
    if usuario.perfil == "sindicato" and row.get("id_sindicato_atual") not in usuario.sindicatos:
        raise HTTPException(403, "Trabalhador fora do escopo do usuário")

    if row.get("titularidade") == "dependente":
        row["dependentes"] = []
        row["titular"] = trabalhador_repo.buscar_titular(row.get("cpf_titular"))
    else:
        row["dependentes"] = trabalhador_repo.buscar_dependentes(row.get("cpf"))
        row["titular"] = None

    return row


@router.get("/{id_trabalhador}")
def detalhe(
    id_trabalhador: int,
    usuario: Annotated[UsuarioInfo, Depends(usuario_logado)],
):
    row = trabalhador_repo.buscar_por_id(id_trabalhador)
    if not row:
        raise HTTPException(404, "Trabalhador não encontrado")

    # RLS: empresa só vê trabalhador das suas empresas; sindicato só do seu sind
    if usuario.perfil == "empresa" and row.get("id_empresa_atual") not in usuario.empresas:
        raise HTTPException(403, "Trabalhador fora do escopo do usuário")
    if usuario.perfil == "sindicato" and row.get("id_sindicato_atual") not in usuario.sindicatos:
        raise HTTPException(403, "Trabalhador fora do escopo do usuário")

    return row
