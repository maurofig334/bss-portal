"""
/contas-pagar — contas a pagar (bss.pagamento). RESTRITO à equipe interna.

GET  /contas-pagar            → lista com filtros + soma
GET  /contas-pagar/exportar   → .xlsx do filtro (o XLS que colam no financeiro)
POST /contas-pagar/liquidar   → marca parcelas como pagas
"""

from datetime import date, datetime
from io import BytesIO
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from .auth import UsuarioInfo, exigir_interno
from . import pagamento_repo


router = APIRouter(prefix="/contas-pagar", tags=["contas-pagar"])


@router.get("")
def listar(
    usuario: Annotated[UsuarioInfo, Depends(exigir_interno)],
    busca: str | None = None,
    status: str | None = None,
    forma: str | None = None,
    empresa_cnpj: str | None = None,
    data_de: str | None = None,
    data_ate: str | None = None,
    pagina: int = 1,
    por_pagina: int = 50,
    ordem: str = "data_prevista",
    desc: bool = True,
):
    return pagamento_repo.listar(
        busca=busca, status=status, forma=forma, empresa_cnpj=empresa_cnpj,
        data_de=data_de, data_ate=data_ate,
        pagina=pagina, por_pagina=por_pagina, ordem=ordem, desc=desc,
    )


@router.get("/exportar")
def exportar(
    usuario: Annotated[UsuarioInfo, Depends(exigir_interno)],
    busca: str | None = None,
    status: str | None = None,
    forma: str | None = None,
    empresa_cnpj: str | None = None,
    data_de: str | None = None,
    data_ate: str | None = None,
    ordem: str = "data_prevista",
    desc: bool = True,
):
    """
    Exporta o filtro em .xlsx — TODAS as linhas, sem paginação.

    Este é o fluxo que a BSS já usava: exportar os pagamentos e colar no
    sistema financeiro externo. ❓ O layout de colunas que aquele sistema
    espera ainda não foi definido — por enquanto exporta um layout genérico e
    completo. Ajustar quando a BSS passar o formato de importação.
    """
    linhas = pagamento_repo.listar_tudo(
        busca=busca, status=status, forma=forma, empresa_cnpj=empresa_cnpj,
        data_de=data_de, data_ate=data_ate, ordem=ordem, desc=desc,
    )

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Contas a Pagar"
    cab = ["Protocolo", "Parcela", "Beneficiário", "CPF", "Valor", "Status",
           "Forma", "Data prevista", "Data pagamento",
           "Empresa", "CNPJ", "Sindicato", "Tipo", "Trabalhador"]
    ws.append(cab)
    bold = Font(bold=True)
    fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    for c in ws[1]:
        c.font = bold
        c.fill = fill

    def _d(v):
        return v if isinstance(v, (date, datetime)) else (v or "")

    for p in linhas:
        ws.append([
            p["protocolo"] or "", p["parcela"], p["beneficiario_nome"] or "",
            p["beneficiario_cpf"] or "", float(p["valor"] or 0), p["status"] or "",
            p["forma_pagamento"] or "", _d(p["data_referencia"]), _d(p["data_pagamento"]),
            p["empresa"] or "", p["empresa_cnpj"] or "", p["sindicato"] or "",
            p["tipo_beneficio"] or "", p["trabalhador"] or "",
        ])

    larg = {"A": 16, "B": 8, "C": 30, "D": 15, "E": 12, "F": 11, "G": 10,
            "H": 13, "I": 14, "J": 34, "K": 20, "L": 34, "M": 16, "N": 30}
    for col, w in larg.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    hoje = date.today().isoformat()
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="contas_a_pagar_{hoje}.xlsx"'},
    )


class LiquidarIn(BaseModel):
    ids: list[int]
    data_pagamento: str | None = None   # ISO; default = hoje


@router.post("/liquidar")
def liquidar(
    dados: LiquidarIn,
    usuario: Annotated[UsuarioInfo, Depends(exigir_interno)],
):
    """
    Marca parcelas como pagas. O BSS passa a ser a fonte da verdade da baixa —
    o SuiteCRM nunca rastreou isso de forma confiável (baixa era dada no
    sistema financeiro externo, por planilha).
    """
    if not dados.ids:
        raise HTTPException(400, "Nenhuma parcela selecionada")

    dt = dados.data_pagamento or date.today().isoformat()
    n = pagamento_repo.liquidar(dados.ids, dt, usuario.id)
    return {"liquidadas": n, "solicitadas": len(dados.ids)}
