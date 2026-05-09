"""
Parser e validador da planilha de Trabalhadores.

Formato esperado (.xlsx):
  Aba "Trabalhadores":
    A: CNPJ DA EMPRESA  (com ou sem pontos/traços)
    B: CPF              (com ou sem pontos/traços; valida DV)
    C: NOME COMPLETO    (não vazio)
    D: SINDICATO LABORAL (texto exato — ver aba LISTA DE SINDICATOS)

  Aba "LISTA DE SINDICATOS" (referência ao usuário): apenas leitura, ignorada
  pelo backend — a fonte de verdade dos sindicatos é bss.sindicato.

Validação puramente de FORMATO neste módulo:
  - CNPJ tem 14 dígitos
  - CPF tem 11 dígitos + DV válido
  - Nome não vazio
  - Sindicato não vazio
  - Não há (CNPJ, CPF) duplicado na mesma planilha

Validações que dependem do banco (CNPJ existe em bss.empresa, sindicato bate
em bss.sindicato.razao_social, usuário tem permissão pro CNPJ) ficam no router
que orquestra a importação.
"""

from __future__ import annotations

import re
from io import BytesIO
from typing import Any, NamedTuple


# ──────────────────────────────────────────────────────────────────────────
# Estruturas de retorno
# ──────────────────────────────────────────────────────────────────────────
class LinhaPlanilha(NamedTuple):
    linha: int          # número da linha no Excel (>=2)
    cnpj: str           # 14 dígitos (sem formatação)
    cpf: str            # 11 dígitos
    nome: str           # nome completo (trim)
    sindicato: str      # texto bruto do sindicato (trim)


class ResultadoParser(NamedTuple):
    linhas: list[LinhaPlanilha]
    erros: dict[str, list[str]]   # categoria → lista de valores


# Categorias de erro (estáveis pra renderizar no front):
CATEGORIAS_ERRO = (
    "estrutura",                    # ['Aba "Trabalhadores" não encontrada', ...]
    "cnpj_invalido",                # ['01.571.635/0002-01', ...]  -- 14 dígitos
    "cpf_invalido",                 # ['40654062', ...]            -- 11 dígitos+DV
    "nome_vazio",                   # ['linha 5', 'linha 7', ...]
    "sindicato_vazio",              # ['linha 5', ...]
    "cpf_duplicado_na_planilha",    # ['12345678901', ...]        -- mesmo CNPJ+CPF
)


# ──────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────
def _so_digitos(s: Any) -> str:
    if s is None:
        return ""
    return re.sub(r"\D+", "", str(s))


def _norm_cpf(s: Any) -> str:
    """Só dígitos, pad-left com zeros até 11 (Excel come zero à esquerda)."""
    d = _so_digitos(s)
    if not d or len(d) > 11:
        return d
    return d.zfill(11)


def _norm_cnpj(s: Any) -> str:
    """Só dígitos, pad-left com zeros até 14."""
    d = _so_digitos(s)
    if not d or len(d) > 14:
        return d
    return d.zfill(14)


def validar_cpf(cpf_raw: Any) -> bool:
    """Valida CPF: 11 dígitos + DV (Mod 11). Rejeita XXXXXXXXXXX (todos iguais).
    Aceita CPF que perdeu zero à esquerda (faz pad antes de validar)."""
    cpf = _norm_cpf(cpf_raw)
    if len(cpf) != 11 or cpf == cpf[0] * 11:
        return False
    # DV1
    soma = sum(int(cpf[i]) * (10 - i) for i in range(9))
    dv1 = (soma * 10) % 11
    if dv1 == 10:
        dv1 = 0
    if int(cpf[9]) != dv1:
        return False
    # DV2
    soma = sum(int(cpf[i]) * (11 - i) for i in range(10))
    dv2 = (soma * 10) % 11
    if dv2 == 10:
        dv2 = 0
    if int(cpf[10]) != dv2:
        return False
    return True


def _str_or_empty(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _norm_nome(v: Any) -> str:
    """Trim + colapsa espaços internos. Excel costuma ter espaço duplo, leading/trailing."""
    s = _str_or_empty(v)
    if not s:
        return ""
    return re.sub(r"\s+", " ", s)


def _linha_vazia(*vals: Any) -> bool:
    return all(_str_or_empty(v) == "" for v in vals)


# ──────────────────────────────────────────────────────────────────────────
# Parser principal
# ──────────────────────────────────────────────────────────────────────────
def parse_planilha_trabalhadores(file_bytes: bytes) -> ResultadoParser:
    """
    Lê o .xlsx e retorna (linhas válidas em formato, erros por categoria).
    NÃO consulta banco — só valida formato.

    Erros retornados são acumulados por categoria, sem repetições.
    """
    erros: dict[str, list[str]] = {c: [] for c in CATEGORIAS_ERRO}

    # Importa lazy pra não pesar startup quando upload não for usado
    try:
        from openpyxl import load_workbook
    except ImportError:
        erros["estrutura"].append("Biblioteca openpyxl ausente no servidor.")
        return ResultadoParser([], erros)

    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        erros["estrutura"].append(f"Arquivo inválido ou corrompido: {e}")
        return ResultadoParser([], erros)

    # Aba certa?
    nome_aba = "Trabalhadores"
    if nome_aba not in wb.sheetnames:
        # Tenta variações comuns (case-insensitive)
        match = next((s for s in wb.sheetnames if s.upper() == nome_aba.upper()), None)
        if match:
            nome_aba = match
        else:
            erros["estrutura"].append(
                f"Aba '{nome_aba}' não encontrada. Abas no arquivo: {wb.sheetnames}"
            )
            return ResultadoParser([], erros)

    ws = wb[nome_aba]

    # Header (linha 1) — confere as 4 primeiras colunas (tolerante a variações)
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        erros["estrutura"].append("Aba 'Trabalhadores' está vazia.")
        return ResultadoParser([], erros)

    esperados = ("CNPJ", "CPF", "NOME", "SINDICATO")
    header_upper = [_str_or_empty(c).upper() for c in (header or ())][:4]
    if len(header_upper) < 4 or not all(
        esp in (header_upper[i] if i < len(header_upper) else "")
        for i, esp in enumerate(esperados)
    ):
        erros["estrutura"].append(
            "Cabeçalho deve conter: CNPJ DA EMPRESA, CPF, NOME COMPLETO, SINDICATO LABORAL"
            f" (encontrado: {header_upper})"
        )
        return ResultadoParser([], erros)

    # Itera linhas de dado (a partir da linha 2)
    linhas: list[LinhaPlanilha] = []
    chaves_vistas: set[tuple[str, str]] = set()
    n_linha = 1

    for raw_row in rows_iter:
        n_linha += 1
        if not raw_row:
            continue
        cnpj_raw = raw_row[0] if len(raw_row) > 0 else None
        cpf_raw  = raw_row[1] if len(raw_row) > 1 else None
        nome_raw = raw_row[2] if len(raw_row) > 2 else None
        sind_raw = raw_row[3] if len(raw_row) > 3 else None

        # Linha totalmente em branco: ignora
        if _linha_vazia(cnpj_raw, cpf_raw, nome_raw, sind_raw):
            continue

        cnpj = _norm_cnpj(cnpj_raw)
        cpf  = _norm_cpf(cpf_raw)
        nome = _norm_nome(nome_raw)
        sind = _str_or_empty(sind_raw)

        # Valida cada campo. Acumula erros mesmo se a linha tiver vários.
        ok = True
        if len(cnpj) != 14:
            erros["cnpj_invalido"].append(_str_or_empty(cnpj_raw) or f"(linha {n_linha})")
            ok = False
        if not validar_cpf(cpf):
            erros["cpf_invalido"].append(_str_or_empty(cpf_raw) or f"(linha {n_linha})")
            ok = False
        if not nome:
            erros["nome_vazio"].append(f"linha {n_linha}")
            ok = False
        if not sind:
            erros["sindicato_vazio"].append(f"linha {n_linha}")
            ok = False

        # Duplicidade (mesmo CNPJ+CPF dentro da planilha)
        if ok and len(cnpj) == 14 and len(cpf) == 11:
            chave = (cnpj, cpf)
            if chave in chaves_vistas:
                erros["cpf_duplicado_na_planilha"].append(cpf)
                ok = False
            else:
                chaves_vistas.add(chave)

        if ok:
            linhas.append(LinhaPlanilha(
                linha=n_linha, cnpj=cnpj, cpf=cpf, nome=nome, sindicato=sind,
            ))

    # Deduplica erros mantendo ordem (set não preserva)
    for cat, lst in erros.items():
        seen = set()
        dedup = []
        for v in lst:
            if v not in seen:
                seen.add(v)
                dedup.append(v)
        erros[cat] = dedup

    return ResultadoParser(linhas, erros)


def total_erros(erros: dict[str, list[str]]) -> int:
    return sum(len(v) for v in erros.values())

# ============================================================================
# PLANILHA DE INATIVAÇÃO
# ============================================================================
# Aba "Planilha1" com 2 colunas: A=CNPJ DA EMPRESA, B=CPF
# Aplica-se POR (CNPJ, CPF): marca o vínculo do trabalhador na empresa como inativo.
# Idempotente — duplicidades dentro da planilha são ignoradas silenciosamente.

CATEGORIAS_ERRO_INATIVACAO = (
    "estrutura",
    "cnpj_invalido",
    "cnpj_nao_cadastrado",
    "cnpj_sem_permissao",
    "cpf_invalido",
    "cpf_nao_cadastrado",
    "cpf_nao_ativo_neste_cnpj",
)


class LinhaInativacao(NamedTuple):
    linha: int
    cnpj: str
    cpf: str


def parse_planilha_inativacao(file_bytes: bytes) -> tuple[list, dict]:
    """Lê o .xlsx de inativação. Retorna (linhas válidas, erros por categoria)."""
    erros: dict[str, list[str]] = {c: [] for c in CATEGORIAS_ERRO_INATIVACAO}

    try:
        from openpyxl import load_workbook
    except ImportError:
        erros["estrutura"].append("Biblioteca openpyxl ausente no servidor.")
        return [], erros

    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        erros["estrutura"].append(f"Arquivo inválido ou corrompido: {e}")
        return [], erros

    # Aceita "Planilha1" ou primeira aba (case-insensitive)
    nome_aba = None
    for cand in ("Planilha1", "Inativacao", "Inativação"):
        match = next((s for s in wb.sheetnames if s.upper() == cand.upper()), None)
        if match:
            nome_aba = match
            break
    if nome_aba is None:
        nome_aba = wb.sheetnames[0]
    ws = wb[nome_aba]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        erros["estrutura"].append("Planilha vazia.")
        return [], erros

    header_upper = [_str_or_empty(c).upper() for c in (header or ())][:2]
    if len(header_upper) < 2 or "CNPJ" not in header_upper[0] or "CPF" not in header_upper[1]:
        erros["estrutura"].append(
            "Cabeçalho deve conter: CNPJ DA EMPRESA, CPF "
            f"(encontrado: {header_upper})"
        )
        return [], erros

    linhas: list[LinhaInativacao] = []
    chaves_vistas: set[tuple[str, str]] = set()
    n_linha = 1
    for raw_row in rows_iter:
        n_linha += 1
        if not raw_row:
            continue
        cnpj_raw = raw_row[0] if len(raw_row) > 0 else None
        cpf_raw = raw_row[1] if len(raw_row) > 1 else None
        if _linha_vazia(cnpj_raw, cpf_raw):
            continue
        cnpj = _norm_cnpj(cnpj_raw)
        cpf = _norm_cpf(cpf_raw)
        if len(cnpj) != 14:
            erros["cnpj_invalido"].append(_str_or_empty(cnpj_raw) or f"(linha {n_linha})")
            continue
        if not validar_cpf(cpf):
            erros["cpf_invalido"].append(_str_or_empty(cpf_raw) or f"(linha {n_linha})")
            continue
        chave = (cnpj, cpf)
        if chave in chaves_vistas:
            continue  # silenciosamente ignora duplicado
        chaves_vistas.add(chave)
        linhas.append(LinhaInativacao(linha=n_linha, cnpj=cnpj, cpf=cpf))

    # dedup erros mantendo ordem
    for cat, lst in erros.items():
        seen = set()
        dedup = []
        for v in lst:
            if v not in seen:
                seen.add(v)
                dedup.append(v)
        erros[cat] = dedup

    return linhas, erros


# ============================================================================
# PLANILHA DE DEPENDENTES
# ============================================================================
# Aba "Dependentes" com 4 colunas:
#   A=CNPJ DA EMPRESA, B=CPF DO DEPENDENTE, C=NOME COMPLETO DO DEPENDENTE,
#   D=CPF DO TRABALHADOR (titular).

CATEGORIAS_ERRO_DEPENDENTES = (
    "estrutura",
    "cnpj_invalido",
    "cnpj_nao_cadastrado",
    "cnpj_sem_permissao",
    "cpf_dependente_invalido",
    "cpf_titular_invalido",
    "titular_nao_cadastrado",
    "titular_nao_ativo_neste_cnpj",
    "nome_dependente_vazio",
    "cpf_dependente_duplicado_na_planilha",
)


class LinhaDependente(NamedTuple):
    linha: int
    cnpj: str
    cpf_dep: str
    nome_dep: str
    cpf_titular: str


def parse_planilha_dependentes(file_bytes: bytes) -> tuple[list, dict]:
    erros: dict[str, list[str]] = {c: [] for c in CATEGORIAS_ERRO_DEPENDENTES}
    try:
        from openpyxl import load_workbook
    except ImportError:
        erros["estrutura"].append("Biblioteca openpyxl ausente no servidor.")
        return [], erros

    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        erros["estrutura"].append(f"Arquivo inválido ou corrompido: {e}")
        return [], erros

    nome_aba = None
    for cand in ("Dependentes", "Planilha1"):
        match = next((s for s in wb.sheetnames if s.upper() == cand.upper()), None)
        if match:
            nome_aba = match
            break
    if nome_aba is None:
        nome_aba = wb.sheetnames[0]
    ws = wb[nome_aba]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        erros["estrutura"].append("Planilha vazia.")
        return [], erros

    header_upper = [_str_or_empty(c).upper() for c in (header or ())][:4]
    esperado = ("CNPJ", "CPF", "NOME", "CPF")  # CPF do titular tb tem "CPF"
    if (len(header_upper) < 4
        or "CNPJ" not in header_upper[0]
        or "CPF" not in header_upper[1]
        or "NOME" not in header_upper[2]
        or "CPF" not in header_upper[3]):
        erros["estrutura"].append(
            "Cabeçalho deve conter: CNPJ DA EMPRESA, CPF DO DEPENDENTE, "
            "NOME COMPLETO DO DEPENDENTE, CPF DO TRABALHADOR"
            f" (encontrado: {header_upper})"
        )
        return [], erros

    linhas: list[LinhaDependente] = []
    cpfs_dep_vistos: set[str] = set()
    n_linha = 1
    for raw_row in rows_iter:
        n_linha += 1
        if not raw_row:
            continue
        cnpj_raw = raw_row[0] if len(raw_row) > 0 else None
        cpf_dep_raw = raw_row[1] if len(raw_row) > 1 else None
        nome_raw = raw_row[2] if len(raw_row) > 2 else None
        cpf_tit_raw = raw_row[3] if len(raw_row) > 3 else None

        if _linha_vazia(cnpj_raw, cpf_dep_raw, nome_raw, cpf_tit_raw):
            continue
        cnpj = _norm_cnpj(cnpj_raw)
        cpf_dep = _norm_cpf(cpf_dep_raw)
        cpf_tit = _norm_cpf(cpf_tit_raw)
        nome = _norm_nome(nome_raw)
        ok = True
        if len(cnpj) != 14:
            erros["cnpj_invalido"].append(_str_or_empty(cnpj_raw) or f"(linha {n_linha})")
            ok = False
        if not validar_cpf(cpf_dep):
            erros["cpf_dependente_invalido"].append(_str_or_empty(cpf_dep_raw) or f"(linha {n_linha})")
            ok = False
        if not validar_cpf(cpf_tit):
            erros["cpf_titular_invalido"].append(_str_or_empty(cpf_tit_raw) or f"(linha {n_linha})")
            ok = False
        if not nome:
            erros["nome_dependente_vazio"].append(f"linha {n_linha}")
            ok = False
        if ok:
            if cpf_dep in cpfs_dep_vistos:
                erros["cpf_dependente_duplicado_na_planilha"].append(cpf_dep)
                ok = False
            else:
                cpfs_dep_vistos.add(cpf_dep)
        if ok:
            linhas.append(LinhaDependente(
                linha=n_linha, cnpj=cnpj, cpf_dep=cpf_dep,
                nome_dep=nome, cpf_titular=cpf_tit,
            ))

    for cat, lst in erros.items():
        seen = set()
        dedup = []
        for v in lst:
            if v not in seen:
                seen.add(v); dedup.append(v)
        erros[cat] = dedup

    return linhas, erros
