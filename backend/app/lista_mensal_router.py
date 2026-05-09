"""
Endpoints de upload de planilha mensal de trabalhadores.

  POST /lista-mensal/upload-trabalhadores  → recebe .xlsx, valida e grava
  GET  /lista-mensal/template               → baixa planilha modelo (.xlsx)
"""

from __future__ import annotations

from io import BytesIO
from typing import Annotated

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse

from .auth import UsuarioInfo, usuario_logado
from .database import get_pg_connection
from .import_planilha import (
    parse_planilha_trabalhadores,
    parse_planilha_inativacao,
    parse_planilha_dependentes,
    total_erros,
    CATEGORIAS_ERRO,
)
from . import lista_mensal_repo as repo


router = APIRouter(prefix="/lista-mensal", tags=["lista-mensal"])


# ──────────────────────────────────────────────────────────────────────────
# UPLOAD
# ──────────────────────────────────────────────────────────────────────────
PERFIS_PERMITIDOS_UPLOAD = {"empresa", "admin", "interno"}
MAX_BYTES = 20 * 1024 * 1024   # 20 MB


@router.post("/upload-trabalhadores")
async def upload_trabalhadores(
    usuario: Annotated[UsuarioInfo, Depends(usuario_logado)],
    arquivo: UploadFile = File(...),
):
    if usuario.perfil not in PERFIS_PERMITIDOS_UPLOAD:
        raise HTTPException(403, "Perfil sem permissão para upload de planilha.")

    nome = arquivo.filename or "trabalhadores.xlsx"
    if not nome.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Arquivo deve ser .xlsx (Excel).")

    conteudo = await arquivo.read()
    if len(conteudo) > MAX_BYTES:
        raise HTTPException(413, f"Arquivo maior que {MAX_BYTES // (1024*1024)} MB.")
    if len(conteudo) == 0:
        raise HTTPException(400, "Arquivo vazio.")

    # ── Fase 1: parsing puro (formato + DV de CPF) ─────────────────────
    parsed = parse_planilha_trabalhadores(conteudo)

    erros = {k: list(v) for k, v in parsed.erros.items()}

    # Se já tem erro de estrutura, aborta sem nem checar banco
    if erros["estrutura"]:
        raise HTTPException(
            400,
            detail={
                "mensagem": "Erros de estrutura no arquivo.",
                "erros": _sanitiza_erros(erros),
            },
        )

    # ── Fase 2: validações contra o banco ──────────────────────────────
    cnpjs = sorted({l.cnpj for l in parsed.linhas})
    sinds = sorted({l.sindicato for l in parsed.linhas})

    map_emp = repo.resolver_empresas(cnpjs)
    map_sind = repo.resolver_sindicatos(sinds)

    # CNPJ não cadastrado em bss.empresa
    cnpjs_invalidos = [c for c in cnpjs if c not in map_emp]
    if cnpjs_invalidos:
        # formata pra exibir mascarado
        erros.setdefault("cnpj_nao_cadastrado", [])
        for c in cnpjs_invalidos:
            erros["cnpj_nao_cadastrado"].append(_fmt_cnpj(c))

    # CNPJs sem permissão (perfil empresa)
    if usuario.perfil == "empresa":
        permitidas = repo.empresas_do_usuario(usuario.id)
        sem_perm = [
            _fmt_cnpj(c)
            for c in cnpjs
            if c in map_emp and map_emp[c] not in permitidas
        ]
        if sem_perm:
            erros.setdefault("cnpj_sem_permissao", [])
            erros["cnpj_sem_permissao"].extend(sem_perm)

    # Sindicatos não encontrados
    chave_sind_map = {repo._norm_sindicato(s): s for s in sinds}
    sinds_invalidos = [
        chave_sind_map[k] for k in chave_sind_map
        if k not in map_sind
    ]
    if sinds_invalidos:
        erros.setdefault("sindicato_nao_cadastrado", [])
        erros["sindicato_nao_cadastrado"].extend(sinds_invalidos)

    # Se algum erro acumulou, aborta SEM gravar nada
    erros_uteis = {k: v for k, v in erros.items() if v}
    if erros_uteis:
        raise HTTPException(
            400,
            detail={
                "mensagem": "Carga não realizada — corrija os erros e refaça o procedimento.",
                "erros": erros_uteis,
            },
        )

    # ── Fase 3: gravação em transação ──────────────────────────────────
    linhas_para_gravar = []
    for ln in parsed.linhas:
        linhas_para_gravar.append({
            "linha":        ln.linha,
            "cnpj":         ln.cnpj,
            "cpf":          ln.cpf,
            "nome":         ln.nome,
            "sindicato":    ln.sindicato,
            "id_empresa":   map_emp[ln.cnpj],
            "id_sindicato": map_sind[repo._norm_sindicato(ln.sindicato)],
        })

    resultado = repo.processar_carga_trabalhadores(
        linhas=linhas_para_gravar,
        mes_referencia=repo.primeiro_dia_do_mes(),
        id_usuario_upload=usuario.id,
        arquivo_nome=nome,
    )
    return {
        "mensagem": "Carga concluída.",
        "qtd_processadas": resultado["qtd_processadas"],
        "qtd_inativadas":  resultado["qtd_inativadas"],
        "qtd_listas_criadas": len(resultado["listas_criadas"]),
        "lista_mensal_ids": resultado["listas_criadas"],
    }


def _sanitiza_erros(d: dict[str, list[str]]) -> dict[str, list[str]]:
    """Remove categorias vazias do payload final."""
    return {k: v for k, v in d.items() if v}


def _fmt_cnpj(c: str) -> str:
    if not c or len(c) != 14:
        return c
    return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}"


# ──────────────────────────────────────────────────────────────────────────
# TEMPLATE (download da planilha modelo)
# ──────────────────────────────────────────────────────────────────────────
@router.get("/template")
def baixar_template(
    usuario: Annotated[UsuarioInfo, Depends(usuario_logado)],
):
    """
    Gera dinamicamente o .xlsx modelo:
      - Aba 'Trabalhadores' com cabeçalho
      - Aba 'LISTA DE SINDICATOS' com nomes vindos de bss.sindicato (ativos,
        em atendimento, ordenados alfabeticamente)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    sql_sindicatos = """
        SELECT razao_social
          FROM bss.sindicato
         WHERE ativo = TRUE AND em_atendimento = TRUE
         ORDER BY razao_social
    """
    with get_pg_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(sql_sindicatos)
            nomes = [r["razao_social"] for r in cur.fetchall()]

    wb = Workbook()
    ws = wb.active
    ws.title = "Trabalhadores"
    ws.append(["CNPJ DA EMPRESA", "CPF", "NOME COMPLETO", "SINDICATO LABORAL", "", "Para saber quais nomes de sindicatos usar, veja na aba LISTA DE SINDICATOS"])

    # Estilo do header
    bold = Font(bold=True)
    fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    for cell in ws[1][:4]:
        cell.font = bold
        cell.fill = fill

    # Larguras
    larguras = {"A": 22, "B": 16, "C": 35, "D": 38, "F": 60}
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w

    # Aba sindicatos
    ws2 = wb.create_sheet("LISTA DE SINDICATOS")
    for n in nomes:
        ws2.append([n])
    ws2.column_dimensions["A"].width = 50

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="trabalhadores_modelo.xlsx"'},
    )

# ──────────────────────────────────────────────────────────────────────────
# INATIVAÇÃO (delta) — fluxo de 2 etapas via query ?confirmar=true|false
# ──────────────────────────────────────────────────────────────────────────
from fastapi import Query


@router.post("/upload-inativacao")
async def upload_inativacao(
    usuario: Annotated[UsuarioInfo, Depends(usuario_logado)],
    arquivo: UploadFile = File(...),
    confirmar: bool = Query(False, description="false=preview; true=executa os válidos"),
):
    if usuario.perfil not in PERFIS_PERMITIDOS_UPLOAD:
        raise HTTPException(403, "Perfil sem permissão para upload de planilha.")

    nome = arquivo.filename or "trabalhadores_inativos.xlsx"
    if not nome.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Arquivo deve ser .xlsx (Excel).")
    conteudo = await arquivo.read()
    if len(conteudo) > MAX_BYTES:
        raise HTTPException(413, f"Arquivo maior que {MAX_BYTES // (1024*1024)} MB.")
    if not conteudo:
        raise HTTPException(400, "Arquivo vazio.")

    # Fase 1: parse formato
    linhas, erros_fmt = parse_planilha_inativacao(conteudo)
    if erros_fmt.get("estrutura"):
        raise HTTPException(400, detail={
            "mensagem": "Erros de estrutura no arquivo.",
            "erros": _sanitiza_erros(erros_fmt),
        })

    # Fase 2: análise contra banco (read-only)
    mes_ref = repo.primeiro_dia_do_mes()
    analise = repo.analisar_inativacao(
        linhas_planilha=linhas,
        id_usuario_upload=usuario.id,
        perfil_usuario=usuario.perfil,
        mes_referencia=mes_ref,
    )
    erros_combinados = {**_sanitiza_erros(erros_fmt), **analise["erros"]}
    erros_combinados = {k: v for k, v in erros_combinados.items() if v}

    if not confirmar:
        # Preview: devolve relatório sem gravar
        return {
            "modo": "preview",
            "mes_referencia": mes_ref.isoformat(),
            "qtd_validas": len(analise["validas"]),
            "qtd_total_planilha": len(linhas) + sum(
                len(v) for v in erros_fmt.values()
            ),
            "erros": erros_combinados,
            "preview_validas": [
                {"cnpj": _fmt_cnpj(v["cnpj"]), "cpf": v["cpf"]}
                for v in analise["validas"][:50]
            ],
        }

    # Fase 3: efetivar (só os válidos)
    if not analise["validas"]:
        raise HTTPException(400, detail={
            "mensagem": "Nenhuma linha válida pra efetivar.",
            "erros": erros_combinados,
        })

    resultado = repo.efetivar_inativacao(
        validas=analise["validas"],
        mes_referencia=mes_ref,
        arquivo_nome=nome,
    )
    return {
        "modo": "executado",
        "mensagem": f"{resultado['qtd_inativadas']} trabalhador(es) inativado(s).",
        "qtd_inativadas": resultado["qtd_inativadas"],
        "qtd_listas_criadas": len(resultado["listas_criadas"]),
        "lista_mensal_ids": resultado["listas_criadas"],
        "erros_ignorados": erros_combinados,
    }


# ──────────────────────────────────────────────────────────────────────────
# DEPENDENTES — all-or-nothing igual o Carregar
# ──────────────────────────────────────────────────────────────────────────
@router.post("/upload-dependentes")
async def upload_dependentes(
    usuario: Annotated[UsuarioInfo, Depends(usuario_logado)],
    arquivo: UploadFile = File(...),
):
    if usuario.perfil not in PERFIS_PERMITIDOS_UPLOAD:
        raise HTTPException(403, "Perfil sem permissão para upload de planilha.")

    nome = arquivo.filename or "dependentes.xlsx"
    if not nome.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Arquivo deve ser .xlsx (Excel).")
    conteudo = await arquivo.read()
    if len(conteudo) > MAX_BYTES:
        raise HTTPException(413, f"Arquivo maior que {MAX_BYTES // (1024*1024)} MB.")
    if not conteudo:
        raise HTTPException(400, "Arquivo vazio.")

    linhas, erros_fmt = parse_planilha_dependentes(conteudo)
    if erros_fmt.get("estrutura"):
        raise HTTPException(400, detail={
            "mensagem": "Erros de estrutura no arquivo.",
            "erros": _sanitiza_erros(erros_fmt),
        })

    erros_fmt_clean = _sanitiza_erros(erros_fmt)
    if erros_fmt_clean:
        raise HTTPException(400, detail={
            "mensagem": "Carga não realizada — corrija os erros e refaça o procedimento.",
            "erros": erros_fmt_clean,
        })

    resultado = repo.analisar_e_processar_dependentes(
        linhas_planilha=linhas,
        id_usuario_upload=usuario.id,
        perfil_usuario=usuario.perfil,
        mes_referencia=repo.primeiro_dia_do_mes(),
        arquivo_nome=nome,
    )

    if resultado["erros"]:
        raise HTTPException(400, detail={
            "mensagem": "Carga não realizada — corrija os erros e refaça o procedimento.",
            "erros": resultado["erros"],
        })

    return {
        "mensagem": "Carga de dependentes concluída.",
        "qtd_processadas": resultado["qtd_processadas"],
        "qtd_listas_criadas": len(resultado["listas_criadas"]),
        "lista_mensal_ids": resultado["listas_criadas"],
    }


# ──────────────────────────────────────────────────────────────────────────
# TEMPLATES (download) — Inativação e Dependentes
# ──────────────────────────────────────────────────────────────────────────
@router.get("/template-inativacao")
def baixar_template_inativacao(
    usuario: Annotated[UsuarioInfo, Depends(usuario_logado)],
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Planilha1"
    ws.append(["CNPJ DA EMPRESA", "CPF"])
    bold = Font(bold=True)
    fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    for c in ws[1]:
        c.font = bold; c.fill = fill
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="trabalhadores_inativos_modelo.xlsx"'},
    )


@router.get("/template-dependentes")
def baixar_template_dependentes(
    usuario: Annotated[UsuarioInfo, Depends(usuario_logado)],
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Dependentes"
    ws.append(["CNPJ DA EMPRESA", "CPF DO DEPENDENTE",
               "NOME COMPLETO DO DEPENDENTE", "CPF DO TRABALHADOR"])
    bold = Font(bold=True)
    fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    for c in ws[1]:
        c.font = bold; c.fill = fill
    larguras = {"A": 22, "B": 18, "C": 38, "D": 18}
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w
    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="dependentes_modelo.xlsx"'},
    )
